from typing import List, Dict
from openpyxl import Workbook

def schemas_to_table(swagger_schemas: List[Dict[str, any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'hello'
    row_number = 1

    for schema_name in swagger_schemas:
        schema = swagger_schemas[schema_name]

        print("==========================-", schema_name,"-=============================")
        
        required_fields = get_required_fields(schema)

        def x(property_name, property):
            nonlocal row_number 
            
            if is_ref(property):
                ref_schema_name = property['$ref'].split("/")[3]
                schema = swagger_schemas[ref_schema_name]
                properties = schema['properties']

                for p in properties:
                    x(property_name + ' -> ' + p, properties[p])
            elif is_array(property):
                x(property_name, property['items'])
            elif is_object(property):
                properties = property['properties']
                for p in properties:
                    x(property_name + ' -> ' + p, properties[p])
            else:
                writer_property(property_name, property, property_name in required_fields)

                write_to_excel(property_name, property, property_name in required_fields, worksheet, row_number)
                
                row_number += 1

        for property_name in schema['properties']:
            property = schema['properties'][property_name]
            x(property_name, property)

    print("=========================================================================")
    workbook.save('./test.xlsx')
        

def get_required_fields(schema: Dict[str, any]) -> List[str]:
    if 'required' in schema:
        print("Required feild: ", schema['required'])
        return schema['required']
    else:
        return []

def is_ref(property: Dict[str, any]) -> bool:
    return '$ref' in property

def has_enum(property: Dict[str, any]) -> bool:
    return 'enum' in property

def is_array(property: Dict[str, any]) -> bool:
    return 'type' in property and property['type'] == 'array'

def is_object(property: Dict[str, any]) -> bool:
    return 'type' in property and property['type'] == 'object'

def add_required_info(property_info: str, required: bool) -> str:
    if required:
        property_info += ', S'
    else:
        property_info += ', N'
    
    return property_info

def add_enum_info(property_info: str, property: Dict[str, any]) -> str:
    if has_enum(property):
        property_info += ', S, [' + ', '.join(property['enum']) + ']'
    else:
        property_info += ', N, NRA'
    
    return property_info

def writer_property(property_name: str, property: Dict[str, any], required: bool) -> None:
    property_info = property_name

    property_info = add_required_info(property_info, required)
    property_info = add_enum_info(property_info, property)

    print(property_info)

def write_to_excel(property_name: str, property: Dict[str, any], required: bool, worksheet, row: int) -> None:
    worksheet.cell(row=row, column=1, value=property_name)
    worksheet.cell(row=row, column=2, value='S' if required else 'N')
    worksheet.cell(row=row, column=3, value='S' if has_enum(property) else 'N')
    worksheet.cell(row=row, column=4, value='[' + ', '.join(property['enum']) + ']' if has_enum(property) else 'NRA')
